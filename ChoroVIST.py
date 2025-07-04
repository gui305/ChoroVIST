import tkinter as tk
from tkinter import messagebox, filedialog
from pathlib import Path
from openpyxl import load_workbook
from PIL import ImageTk, Image
from choroidalyze import Choroidalyzer
import os
from openpyxl import Workbook
from oct_converter.readers import E2E
import torch
import torchvision.transforms as T
import numpy as np
import traceback
from tkcalendar import DateEntry
import sys



if getattr(sys, 'frozen', False):
    # Executável PyInstaller
    base_path = sys._MEIPASS
else:
    # Código normal em .py
    base_path = os.path.abspath(".")

Path("Datafiles").mkdir(exist_ok=True)
# Initialize the app




file_path = "ChoroVIST_Database.xlsx"

def create_excel_if_not_exists():
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Patient's name",
            "Healthcare number",
            "Eye",
            "Date of analysis",
            "Date of acquisition",
            "Intervention Status",
            "Thickness (µm)",
            "Area ( mm²)",
            "Vascular Index",
            "Vessel Area ( mm²)",           
            "Report"
        ])
        workbook.save(file_path)

# Run this once at startup
create_excel_if_not_exists()

# Function to handle Excel file
def handle_excel(mode='r'):
        if os.path.exists(file_path):
            try:
                workbook = load_workbook(file_path)
                return workbook
            except PermissionError:
                messagebox.showerror("Error", "The Excel file is open. Please close the file and try again.")
                return None
        else:
            messagebox.showerror("Error", "Excel File doesn't exist")
            return None


def load_e2e_volume(e2e_path):
    e2e = E2E(str(e2e_path))
    
    required_attrs = ['patient_id', 'first_name', 'surname', 'sex', 'birthdate', 'acquisition_date']
    for attr in required_attrs:
        if not hasattr(e2e, attr):
            setattr(e2e, attr, "unknown")
    oct_volumes = e2e.read_oct_volume()
    if not oct_volumes or not hasattr(oct_volumes[0], 'volume') or not oct_volumes[0].volume:
        raise ValueError("OCT data is incomplete or corrupted (missing image volume).")

    # Garantir pixel_spacing válido
    pixel_spacing = getattr(oct_volumes[0], 'pixel_spacing', None)
    if not pixel_spacing or not isinstance(pixel_spacing, (list, tuple)) or len(pixel_spacing) < 2:
        raise ValueError("OCT data is incomplete or corrupted (missing pixel spacing).")

    return oct_volumes

def perform_search(app, search_term, right_frame):
    if not search_term.isdigit() or len(search_term) != 9:
        messagebox.showerror("Error", "The healthcare number must have 9 digits.")
        return

    workbook = handle_excel('r')
    if workbook is None:
        return

    sheet = workbook.active
    results = [row for row in sheet.iter_rows(values_only=True) if row[1] == search_term]
    workbook.close()

    if not results:
        messagebox.showinfo("Attention", "No patient found!")
        return

    results = sorted(results, key=lambda r: r[3], reverse=True)
    current_index = [0]

    def show_result(index):
        result = results[index]
        name, health_number, eye, date, exam_date, surgery_status, thickness, area, vascular_index, vessel_area = result[:10]

        info_text = (
    f"Patient's name: {name}\n"
    f"Healthcare number: {health_number}\n"
    f"Data: {date}\n"
    f"Eye: {eye}\n"   
)

        if exam_date:
            info_text += f"Date of acquisition: {exam_date}"
        if surgery_status:
            info_text += f"\nIntervention Status: {surgery_status}"
        info_text +=f"\nThickness: {thickness} µm\n"
        info_text +=f"\nArea: {area} mm²\n"
        info_text +=f"\nVascular Index: {vascular_index}\n"
        info_text +=f"\nVessel Area: {vessel_area} mm²"
        info_text = ''.join(info_text)

        img_dirs = sorted(Path("Datafiles").glob(f"{health_number}*"), reverse=True)
        if img_dirs:
            img_dir = img_dirs[index]
            try:
                images_dict = {
                    "original": img_dir / "original.png",
                    "region": img_dir / "region.png",
                    "vessel": img_dir / "vessel.png",
                    "fovea": img_dir / "fovea.png"
                }
                report_path = img_dir / "report.pdf"
                right_frame.update_content(info_text, images_dict, report_path, page_label_text=f"{index+1} / {len(results)}")
            except Exception as e:
                traceback.print_exc()
                right_frame.update_content(info_text, {}, None, error=f"Error loading images: {e}", page_label_text=f"{index+1} / {len(results)}")
        else:
            right_frame.update_content(info_text, {}, None, error="Images not found", page_label_text=f"{index+1} / {len(results)}")

        # Liga os botões de navegação
        right_frame.prev_btn.config(command=lambda: change_page(-1))
        right_frame.next_btn.config(command=lambda: change_page(1))

    def change_page(delta):
        current_index[0] = (current_index[0] + delta) % len(results)
        show_result(current_index[0])

    show_result(current_index[0])

def analyze_and_save(app, patient_name, health_number, eye, oct_path, exam_date, surgery_status, right_frame, insert_frame):
    from datetime import datetime
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Validate patient name
    if not patient_name.strip():
        messagebox.showerror("Error", "Patient's name box can't be empty")
        return
    if not all(c.isalpha() or c.isspace() for c in patient_name):
        messagebox.showerror("Error", "Patient's name must contain only letters and spaces")
        return
    if not health_number.isdigit() or len(health_number) != 9:
        messagebox.showerror("Error", "The healthcare number must have 9 digits")
        return

    # Validate name in Excel
    workbook = handle_excel('r')
    if workbook:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] == health_number:
                db_name = row[0].strip().lower()
                input_name = patient_name.strip().lower()
                if db_name != input_name:
                    messagebox.showerror(
                        "Error",
                        f"The entered name '{patient_name}' does not match the registered name '{row[0]}'. "
                        f"The name field has been corrected to the patient's registered name."
                    )
                    workbook.close()
                    insert_frame.name_entry.delete(0, tk.END)
                    insert_frame.name_entry.insert(0, row[0])
                    return
                break
        workbook.close()

    if eye not in ["Right", "Left"]:
        messagebox.showerror("Error", "Please select the eye (Right or Left).")
        return
    if not oct_path or not Path(oct_path).is_file():
        messagebox.showerror("Error", "Please select a valid file.")
        return

    path = Path(oct_path)
    is_e2e = path.suffix.lower() == ".e2e"

    if is_e2e:
        try:
            oct_volumes = load_e2e_volume(path)
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Critical error", f"Failed to load .e2e:\n\n{e}")
            return
        if not oct_volumes:
            messagebox.showerror("Error", "No valid OCT volumes in .e2e")
            return
        try:

        # Check if there's at least one valid volume with image data
            if not oct_volumes or not hasattr(oct_volumes[0], 'volume') or not oct_volumes[0].volume:
                raise ValueError("OCT data is incomplete or corrupted (missing image volume).")

            oct_volume = oct_volumes[0].volume  # Extract volume

        # Try to get acquisition date, but don't fail if missing/corrupt
            try:
                extracted_exam_date = getattr(oct_volumes[0], 'acquisition_date', None)
                if extracted_exam_date:
                # If extracted_exam_date is datetime object
                    exam_date = (
                    extracted_exam_date.strftime("%Y-%m-%d %H:%M:%S")
                    if hasattr(extracted_exam_date, 'strftime')
                    else str(extracted_exam_date)
                    )
            except Exception as e:
                print(f"Warning: acquisition date missing or corrupted: {e}")
            # Proceed anyway, no need to set exam_date


        # Continue with the analysis pipeline
            try:
                middle_idx = len(oct_volume) // 2
                oct_image = oct_volume[middle_idx]
                oct_image = torch.from_numpy(oct_image).float().unsqueeze(0)

                new_h, new_w = 512, 512
                orig_h, orig_w = oct_image.shape[-2:]
                factor_x = orig_w / new_w
                factor_y = orig_h / new_h
                resize = T.Resize((new_h, new_w))
                oct_image = resize(oct_image)

                pixel_spacing = oct_volumes[0].pixel_spacing
                scale_x = round(pixel_spacing[0] * 1000, 3)
                scale_y = round(pixel_spacing[1] * 1000, 3)
                new_spacing_x = factor_x * scale_x
                new_spacing_y = factor_y * scale_y

                choroidalyzer = Choroidalyzer(default_scale=(new_spacing_x, new_spacing_y))
            except Exception as e:
                traceback.print_exc()
                messagebox.showerror("Critical error", f"Error processing the image volume:\n\n{str(e)}")
                return
            
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror(
            "Critical error",
            f"Failed to process the .e2e file:\n\n{str(e)}"
            )
            return
    else:
        oct_image = oct_path
        choroidalyzer = Choroidalyzer()

    # Try analysis pipeline
    try:
        metrics = choroidalyzer.analyze(oct_image)
        for key in ['thickness', 'area', 'vascular_index', 'vessel_area']:
            value = metrics.get(key)
            if value is None or not isinstance(value, (int, float)) or np.isnan(value) or np.isinf(value):
                raise ValueError(f"Invalid metric value: {key} = {value}")

        metrics = {k: round(metrics[k], 3) for k in metrics if k != 'raw_thickness'}
        metrics['raw_thickness'] = metrics.get('raw_thickness')

        save_dir = Path("Datafiles") / f"{health_number}_{timestamp}"
        info_text = choroidalyzer.predict_and_give_images(
            oct_image, patient_name, health_number, date, eye,
            metrics, exam_date, surgery_status, save_dir
        )

        # Excel update
        workbook = handle_excel('w')
        sheet = workbook.active
        sheet.append([
            patient_name,
            health_number,
            eye,
            date,
            exam_date or "",
            surgery_status or "",
            metrics['thickness'],
            metrics['area'],
            metrics['vascular_index'],
            metrics['vessel_area'],
            
            
        ])
        link_cell = sheet.cell(row=sheet.max_row, column=11)
        link_cell.value = "View Report"

        # Caminho absoluto e no formato file:///
        abs_path = Path(save_dir / 'report.pdf').resolve()
        link_cell.hyperlink = abs_path.as_uri()
        link_cell.style = "Hyperlink"
        workbook.save(file_path)
        workbook.close()

        perform_search(app, health_number, right_frame)
        messagebox.showinfo("Success", "Analysis completed successfully and stored in the database.")

    except Exception as e:
        traceback.print_exc()
        messagebox.showerror("Critical error", f"Failing to process analysis:\n\n{str(e)}")

class ChoroidApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ChoroVIST")
        self.iconbitmap(os.path.join(base_path, "coroide.ico"))
        self.state('zoomed')  # Fullscreen

        self.bemvindo = True  # Flag para saber se ainda está no ecrã de boas-vindas

        # Layout principal fixo
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Fixar tamanhos dos frames
        self.left_frame_width = 300
        self.right_frame_min_width = 900  # Ajusta conforme necessário

        # Menu Frame
        self.menu_frame = MenuFrame(self)
        self.menu_frame.grid(row=0, column=0, columnspan=2, sticky="ew")

        # Left Frame
        self.left_frame = tk.Frame(self, width=self.left_frame_width, bg="lightgray")
        self.left_frame.grid(row=1, column=0, sticky="ns")
        self.left_frame.grid_propagate(False)

        # Right Frame inicial (vazio)
        self.right_frame = tk.Frame(self, bg="white", width=self.right_frame_min_width)
        self.right_frame.grid(row=1, column=1, sticky="nsew")
        self.right_frame.grid_propagate(False)

        self.show_welcome_moderno()

    def show_welcome(self):
        self._clear_frames()
        welcome_label = tk.Label(self.right_frame, text="ChoroVIST",
                                 font=("Helvetica", 16), bg="white")
        welcome_label.pack(padx=20, pady=20)

    def show_welcome_moderno(self):

        self._clear_frames()

        # Carregar logotipos
        try:
            logo_faculdade = Image.open(os.path.join(base_path, 'Tecnico.png')).resize((280, 180))
            self.logo_faculdade_img = ImageTk.PhotoImage(logo_faculdade)
        except:
            self.logo_faculdade_img = None

        try:
            logo_hospital = Image.open(os.path.join(base_path, 'FMUL.png')).resize((200, 200))
            self.logo_hospital_img = ImageTk.PhotoImage(logo_hospital)
        except:
            self.logo_hospital_img = None

        try:
            logo_app = Image.open(os.path.join(base_path, 'OLHO.png')).resize((230, 230))
            self.logo_app_img = ImageTk.PhotoImage(logo_app)
        except:
            self.logo_app_img = None

        # Logotipo da faculdade (esquerda)
        if self.logo_faculdade_img:
            faculdade_label = tk.Label(self.right_frame, image=self.logo_faculdade_img, bg="white")
            faculdade_label.place(x=40, y=20)

        # Logotipo do hospital (direita)
        if self.logo_hospital_img:
            hospital_label = tk.Label(self.right_frame, image=self.logo_hospital_img, bg="white")
            hospital_label.place(relx=1.0, x=-40, y=20, anchor="ne")

        # Título
        title_font = ("Montserrat", 36, "bold")
        title_label = tk.Label(self.right_frame, text="ChoroVIST", font=title_font, bg="white")
        title_label.place(relx=0.5, rely=0.3, anchor="center")

        # Logo da app (opcional)
        if self.logo_app_img:
            logo_app_label = tk.Label(self.right_frame, image=self.logo_app_img, bg="white")
            logo_app_label.place(relx=0.5, rely=0.5, anchor="center")

        # Créditos
        credits_font = ("Open Sans", 12)
        credits_text = (
            "Created by: Guilherme de Matos Trindade Coelho (guilhermemtcoelho@tecnico.ulisboa.pt)\n"
            "Model backend code from doi: https://doi.org/10.1167/iovs.65.6.6\n"
            "Interface, data pipeline, and database structure independently developed.\n"
            "© 2025 Guilherme de Matos Trindade Coelho. All rights reserved"
        )
        credits_label = tk.Label(
            self.right_frame,
            text=credits_text,
            font=credits_font,
            fg="#555555",
            bg="white",
            justify="right"
        )
        credits_label.place(relx=1.0, rely=1.0, x=-20, y=-20, anchor="se")

    def show_insert_screen(self):
        self._clear_frames()

        self.right_frame.destroy()
        self.right_frame = RightFrame(self)
        self.right_frame.grid(row=1, column=1, sticky="nsew")
        self.right_frame.grid_propagate(False)

        InsertFrame(self.left_frame, self.right_frame)

    def show_search_screen(self):
        self._clear_frames()

        self.right_frame.destroy()
        self.right_frame = RightFrame(self)
        self.right_frame.grid(row=1, column=1, sticky="nsew")
        self.right_frame.grid_propagate(False)

        SearchFrame(self.left_frame, self.right_frame)


    def _clear_frames(self):
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

class MenuFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="darkgray")
        insert_btn = tk.Button(self, text="Insert new exam", command=master.show_insert_screen)
        insert_btn.pack(side="left", padx=10, pady=10)

        search_btn = tk.Button(self, text="Search Patient", command=master.show_search_screen)
        search_btn.pack(side="left", padx=10, pady=10)


class InsertFrame(tk.Frame):
    def __init__(self, master, right_frame):
        super().__init__(master, bg="lightgray", width=300)
        self.pack(fill="both", expand=True)
        self.pack_propagate(False)

        tk.Label(self, text="Patient's name:").pack(pady=5)
        self.name_entry = tk.Entry(self)
        self.name_entry.pack(pady=5)

        tk.Label(self, text="Healthcare number:").pack(pady=5)
        self.number_entry = tk.Entry(self)
        self.number_entry.pack(pady=5)
        self.number_entry.bind("<KeyPress>", self.check_number_length)
        self.max_digits_label = tk.Label(
            self,
            text="Healthcare number must contain 9 digits",
            fg="red",
            font=("TkDefaultFont", 10, "bold"),
            bg="lightgray"
        )
        self.max_digits_label.pack_forget()

        tk.Label(self, text="Eye:").pack(pady=5)
        self.eye_var = tk.StringVar()
        tk.Radiobutton(self, text="Left", variable=self.eye_var, value="Left").pack()
        tk.Radiobutton(self, text="Right", variable=self.eye_var, value="Right").pack()
        self.eye_var.set(None)

        tk.Label(self, text="Acquisition Date (optional):\n(.E2E file auto-fills)").pack(pady=5)
        self.exam_date_entry = DateEntry(self, date_pattern="yyyy-mm-dd", showweeknumbers=False)
        self.exam_date_entry.pack(pady=5)

        tk.Label(self, text="Intervention Status (optional):").pack(pady=5)
        self.surgery_status_var = tk.StringVar()
        self.surgery_status_var.set("None")
        surgery_options = ["None", "Pre Intervention", "Post Intervention"]
        tk.OptionMenu(self, self.surgery_status_var, *surgery_options).pack(pady=5)

        tk.Label(self, text="OCT Exam:").pack(pady=5)
        self.oct_image_path_entry = tk.Entry(self)
        self.oct_image_path_entry.pack(pady=5)
        tk.Button(self, text="Browse", command=self.browse_file).pack(pady=5)

        tk.Button(
            self,
            text="Analyze",
            font=("TkDefaultFont", 10, "bold"),
            command=lambda: analyze_and_save(
                self.master.master,
                self.name_entry.get(),
                self.number_entry.get(),
                self.eye_var.get(),
                self.oct_image_path_entry.get(),
                self.exam_date_entry.get(),
                self.surgery_status_var.get(),
                self.right_frame,
                self
            )
        ).pack(pady=10)

        self.right_frame = right_frame

    def browse_file(self):
        file_path = filedialog.askopenfilename()
        if file_path:
            self.oct_image_path_entry.delete(0, tk.END)
            self.oct_image_path_entry.insert(0, file_path)

    def check_number_length(self, event):
        # Allow control keys like backspace
        if event.keysym in ("BackSpace", "Delete", "Left", "Right", "Tab"):
            return

        current_text = self.number_entry.get()
        # If already 9 digits, show warning and block
        if len(current_text) >= 9:
            self.show_warning()
            return "break"  # Prevent the extra digit from being added

    def show_warning(self):
        self.max_digits_label.pack(pady=2)
        self.after(1500, self.max_digits_label.pack_forget)



class SearchFrame(tk.Frame):
    def __init__(self, master, right_frame):
        super().__init__(master, bg="lightgray", width=300)
        self.pack(fill="both", expand=True)
        self.pack_propagate(False)

        tk.Label(self, text="Healthcare number:").pack(pady=5)
        self.number_entry = tk.Entry(self)
        self.number_entry.pack(pady=5)
        self.number_entry.bind("<KeyPress>", self.check_number_length)
        self.number_entry.bind("<KeyRelease>", self.on_keyrelease)

        self.max_digits_label = tk.Label(
            self,
            text="Healthcare number must contain 9 digits",
            fg="red",
            font=("TkDefaultFont", 10, "bold"),
            bg="lightgray"
        )
        self.max_digits_label.pack_forget()

        self.search_button = tk.Button(
            self, text="Search", command=self.on_search_button
        )
        self.search_button.pack(pady=10)

        self.status_label = tk.Label(self, text="", bg="lightgray", fg="black")
        self.status_label.pack(pady=5)

        self.suggestions_frame = None
        self.right_frame = right_frame

    def check_number_length(self, event):
        if event.keysym in ("BackSpace", "Delete", "Left", "Right", "Tab"):
            return

        current_text = self.number_entry.get()
        if len(current_text) >= 9:
            self.show_warning()
            return "break"  # Prevent typing the 10th digit

    def show_warning(self):
        self.max_digits_label.pack(pady=2)
        self.after(1500, self.max_digits_label.pack_forget)

    def on_keyrelease(self, event):
        text = self.number_entry.get()
        if text == "":
            self.clear_suggestions()
            self.number_entry.config(bg="white")
            self.status_label.config(text="")
            return

        if len(text) == 9 and text.isdigit():
            if self.check_in_database(text):
                self.number_entry.config(bg="lightgreen")
                self.status_label.config(
                    text="Patient in database", fg="green"
                )
            else:
                self.number_entry.config(bg="lightcoral")
                self.status_label.config(
                    text="Patient not found in database", fg="red"
                )
            self.clear_suggestions()
            return

        self.number_entry.config(bg="white")
        self.status_label.config(text="")
        suggestions = self.search_database(text)
        self.show_suggestions(suggestions)

    # (Rest of the class remains the same)


    def check_in_database(self, health_number):
        workbook = handle_excel('r')
        if not workbook:
            return False
        sheet = workbook.active
        found = False
        for row in sheet.iter_rows(values_only=True):
            if row[1] and str(row[1]) == health_number:
                found = True
                break
        workbook.close()
        return found

    def search_database(self, query):
        results = []
        seen_numbers = set()
        workbook = handle_excel('r')
        if not workbook:
            return results

        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[1] and str(row[1]).startswith(query):
                health_number = str(row[1])
                if health_number not in seen_numbers:
                    results.append({'number': health_number, 'name': row[0]})
                    seen_numbers.add(health_number)
                if len(results) >= 10:
                    break

        workbook.close()
        return results


    def show_suggestions(self, suggestions):
        self.clear_suggestions()
        if not suggestions:
            return

        self.suggestions_frame = tk.Frame(self, bg="lightgray")
        self.suggestions_frame.pack(pady=5)

        canvas = tk.Canvas(self.suggestions_frame, bg="lightgray", height=200)
        scrollbar = tk.Scrollbar(self.suggestions_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg="lightgray")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for record in suggestions:
            text = f"{record['number']} ({record['name']})"
            btn = tk.Button(
                scrollable_frame, text=text,
                command=lambda r=record: self.select_suggestion(r)
            )
            btn.pack(fill="x", padx=5, pady=2)

    def clear_suggestions(self):
        if self.suggestions_frame:
            self.suggestions_frame.destroy()
            self.suggestions_frame = None

    def select_suggestion(self, record):
        self.number_entry.delete(0, tk.END)
        self.number_entry.insert(0, record['number'])
        self.number_entry.config(bg="lightgreen")
        self.status_label.config(
            text="Patient in database", fg="green"
        )
        self.clear_suggestions()

    def on_search_button(self):
        health_number = self.number_entry.get()
        if len(health_number) != 9 or not health_number.isdigit():
            messagebox.showerror("Error", "Please enter a valid healthcare number (9 digits).")
            return

        # Forçar a verificação da base de dados
        if self.check_in_database(health_number):
            self.number_entry.config(bg="white")
            self.status_label.config(text="")
            perform_search(
        self.master.master,
        self.number_entry.get(),
        self.right_frame
        )
            self.number_entry.delete(0, tk.END)
        else:
            self.number_entry.config(bg="lightcoral")
            self.status_label.config(
                text="Patient not found in database", fg="red"
            )
            perform_search(
        self.master.master,
        self.number_entry.get(),
        self.right_frame
        )
        

class ZoomableImageLabel(tk.Label):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.image_original = None
        self.photo = None
        self.zoom_factor = 1.0
        self.offset_x = 0
        self.offset_y = 0
        self.initial_zoom = 1.0

        self.bind("<MouseWheel>", self.zoom)
        self.bind("<ButtonPress-1>", self.start_pan)
        self.bind("<B1-Motion>", self.pan)
        self.bind("<Configure>", self.reset_offset)

    def set_image(self, pil_image):
        self.image_original = pil_image
        self.offset_x = 0
        self.offset_y = 0

        w, h = self.image_original.size
        label_w, label_h = 300, 300

        scale_w = label_w / w
        scale_h = label_h / h
        self.initial_zoom = min(scale_w, scale_h)

        self.zoom_factor = 1.0
        self.update_image()

    def zoom(self, event):
        factor = 1.1 if event.delta > 0 else 0.9
        new_zoom = self.zoom_factor * factor

        if new_zoom < 1.0:
            new_zoom = 1.0
            self.offset_x = 0
            self.offset_y = 0
        else:
            label_w, label_h = 300, 300
            effective_zoom = self.initial_zoom * self.zoom_factor
            new_effective_zoom = self.initial_zoom * new_zoom

            # Calcular a posição do mouse relativamente à imagem atual
            mouse_x = event.x
            mouse_y = event.y

            img_x = (mouse_x + self.offset_x)
            img_y = (mouse_y + self.offset_y)

            rel_x = img_x / effective_zoom
            rel_y = img_y / effective_zoom

            # Calcular novo offset
            self.offset_x = int(rel_x * new_effective_zoom - mouse_x)
            self.offset_y = int(rel_y * new_effective_zoom - mouse_y)

        self.zoom_factor = new_zoom
        self.update_image()

    def start_pan(self, event):
        self.last_x = event.x
        self.last_y = event.y

    def pan(self, event):
        if self.zoom_factor <= 1.0:
            return
        dx = event.x - self.last_x
        dy = event.y - self.last_y
        self.offset_x -= dx
        self.offset_y -= dy
        self.last_x = event.x
        self.last_y = event.y
        self.update_image()

    def reset_offset(self, event):
        if self.zoom_factor == 1.0:
            self.offset_x = 0
            self.offset_y = 0
            self.update_image()

    def update_image(self):
        if self.image_original:
            w, h = self.image_original.size
            label_w, label_h = 300, 300

            effective_zoom = self.initial_zoom * self.zoom_factor
            new_w = int(w * effective_zoom)
            new_h = int(h * effective_zoom)
            img_resized = self.image_original.resize((new_w, new_h), Image.LANCZOS)

            # Ajustar os offsets (para permitir pan completo)
            max_offset_x = max(0, new_w - label_w)
            max_offset_y = max(0, new_h - label_h)
            self.offset_x = max(-max_offset_x, min(self.offset_x, max_offset_x))
            self.offset_y = max(-max_offset_y, min(self.offset_y, max_offset_y))

            # Calcular a área a exibir
            left = max(0, self.offset_x)
            upper = max(0, self.offset_y)
            right = min(new_w, left + label_w)
            lower = min(new_h, upper + label_h)

            img_cropped = img_resized.crop((left, upper, right, lower))

            self.photo = ImageTk.PhotoImage(img_cropped)
            self.config(image=self.photo, width=300, height=300)
            self.image = self.photo

import textwrap




class RightFrame(tk.Frame):
    def __init__(self, master):
        super().__init__(master, bg="white")
        self.configure(bg="white")

        # Definir a grade com largura fixa para a coluna 0
        self.grid_columnconfigure(0, minsize=350, weight=0)  # Fixa a largura da coluna esquerda
        self.grid_columnconfigure(1, weight=1)               # A coluna direita expande

        # LEFT PANEL — fixed width
        self.left_panel = tk.Frame(self, bg="#f5f5f5", width=350)
        self.left_panel.grid(row=0, column=0, sticky="ns", padx=20, pady=20)
        self.left_panel.grid_propagate(False)  # impede autoajuste dinâmico
        self.left_panel.pack_propagate(False)  # também garante que pack não altere

        # Info box
        self.patient_info_label = tk.Label(
            self.left_panel,
            text="Patient Metadata:\n\n\n\n\n\n",
            justify="left",
            font=("arialbd.ttf", 12),
            bg="#f5f5f5",
            anchor="nw",
            wraplength=320
        )
        self.patient_info_label.pack(pady=(0, 10), anchor="nw", fill="x")

        # Container for View Report and Navigation
        report_nav_frame = tk.Frame(self.left_panel, bg="#f5f5f5")
        report_nav_frame.pack(anchor="nw", pady=(0, 10), fill="x")

        # View Report button inside the new container
        self.report_btn = tk.Button(report_nav_frame, text="View Report", state=tk.DISABLED)
        self.report_btn.pack(side="left", padx=5)

        # Navigation inside the same container
        nav_frame = tk.Frame(report_nav_frame, bg="#f5f5f5")
        nav_frame.pack(side="left", padx=5)

        self.prev_btn = tk.Button(nav_frame, text="<")
        self.prev_btn.pack(side="left", padx=5)
        self.page_label = tk.Label(nav_frame, text="Exam 1 of 1", bg="#f5f5f5")
        self.page_label.pack(side="left", padx=5)
        self.next_btn = tk.Button(nav_frame, text=">")
        self.next_btn.pack(side="left", padx=5)


        # Metrics
        self.metrics_label = tk.Label(
            self.left_panel,
            text="Choroid Metrics:\n\n\n\n",
            justify="left",
            font=("arialbd.ttf", 12),
            bg="#f5f5f5",
            anchor="nw",
            wraplength=320
        )
        self.metrics_label.pack(pady=(0, 10), anchor="nw", fill="x")

        # RIGHT PANEL — images
        self.image_panel = tk.Frame(self, bg="white")
        self.image_panel.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.image_panel.columnconfigure((0, 1), weight=1)
        self.image_panel.rowconfigure((0, 1), weight=1)

        self.blank_thumbnail = self.create_blank_thumbnail(300, 300)

        self.image_labels = {}
        labels = ["Original", "Region", "Vessel", "Fovea"]
        for i, key in enumerate(["original", "region", "vessel", "fovea"]):
            frame = tk.Frame(self.image_panel, width=300, height=330, bg="white")
            frame.grid(row=i // 2, column=i % 2, padx=20, pady=20, sticky="n")
            frame.grid_propagate(False)

            label_title = tk.Label(frame, text=labels[i], font=("arialbd.ttf", 12, "bold"), bg="white")
            label_title.pack(pady=(0, 5))

            image_frame = tk.Frame(frame, width=300, height=300, bg="gray", relief="solid", borderwidth=0)
            image_frame.pack()
            image_frame.pack_propagate(False)

            label = ZoomableImageLabel(image_frame, bg="gray")
            label.config(width=300, height=300)
            label.place(x=0, y=0, width=300, height=300)
            label.set_image(self.blank_thumbnail)

            self.image_labels[key] = label

    def create_blank_thumbnail(self, width, height):
        img = Image.new("RGB", (width, height), color="lightgray")
        return img

    def update_content(self, info_text, images_dict, report_path, error=None, page_label_text=""):
        lines = info_text.split('\n')
        meta_lines = ['Patient Metadata:\n']
        metric_lines = ['Choroid Metrics:\n']
        metrics_keywords = ['Thickness', 'Area', 'Index']

        found_metrics = False
        for line in lines:
            if any(keyword in line for keyword in metrics_keywords):
                found_metrics = True
            if found_metrics:
                metric_lines.append(line)
                found_metrics = False
            else:
                # Wrap long lines nicely at word boundaries (e.g., name)
                wrapped = textwrap.fill(
                    line,
                    width=50,
                    break_long_words=False,
                    break_on_hyphens=False
                )
                meta_lines.append(wrapped)

        self.patient_info_label.config(text="\n".join(meta_lines))
        self.metrics_label.config(text="\n".join(metric_lines))

        for widget in self.left_panel.winfo_children():
            if isinstance(widget, tk.Label) and "Error" in widget.cget("text"):
                widget.destroy()

        for key, label in self.image_labels.items():
            img_path = images_dict.get(key)
            if img_path and os.path.exists(img_path):
                try:
                    img = Image.open(img_path)
                    label.set_image(img)
                except Exception as e:
                    label.set_image(self.blank_thumbnail)
                    traceback.print_exc()
            else:
                label.set_image(self.blank_thumbnail)

        if report_path and os.path.exists(report_path):
            self.report_btn.config(command=lambda: os.startfile(report_path), state=tk.NORMAL)
        else:
            self.report_btn.config(command=lambda: None, state=tk.DISABLED)

        self.page_label.config(text=page_label_text)

        if error:
            error_label = tk.Label(self.left_panel, text=f"Error: {error}", fg="red", bg="#f5f5f5")
            error_label.pack(anchor="nw", pady=(10, 0))

if __name__ == "__main__":
    app = ChoroidApp()
    app.mainloop()
